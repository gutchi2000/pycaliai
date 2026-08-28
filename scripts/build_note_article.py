#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""note 有料記事ドラフト生成 (会場パック / 重賞単体 / 全場パック)。

データ源は `site/data/{date}.json` = build_site.py が `scrub_public()` を通した
公開用ペイロード。記事側でも同じ除外リストを再適用したうえで、記事固有の
追加スクラブ (SHAP `why` の外部指数/調教/計時系、オッズ生値を含む定型文) を掛ける。

  python scripts/build_note_article.py 20260802

出力: reports/note/{date}/{会場}.md, {重賞名}.md, all.md
      reports/note/{date}/_compliance_report.txt
"""
from __future__ import annotations

import base64
import collections
import datetime
import glob
import io
import json
import math
import os
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import build_site  # scrub_public を共有 (JRA-VAN 投稿ガイドライン準拠の単一の真実)

SITE_DATA = ROOT / "site" / "data"
OUT_ROOT = ROOT / "reports" / "note"

PLACE_SLUG = {"札幌": "sapporo", "函館": "hakodate", "福島": "fukushima",
              "新潟": "niigata", "東京": "tokyo", "中山": "nakayama",
              "中京": "chukyo", "京都": "kyoto", "阪神": "hanshin",
              "小倉": "kokura"}

# ------------------------------------------------------------------ 記事側の追加スクラブ
# scrub_public() は horses[].training / taiju / odds / history.runs[].agari3f を落とすが、
# SHAP `why` は素通しなので、外部指数 (TARGET 補正タイム/PCI)・調教タイム・計時 (上り3F/着差)
# 由来の特徴はここで落とす。数値 (`value`) は許可特徴でも一切出力しない。
BLOCKED_WHY = re.compile(
    r"(hosei|trn[HW]|agari|PCI|着差|上り3F|上がり3F|タイム|time|走破|taiju|馬体重|odds|オッズ)",
    re.IGNORECASE)

# 生成 md に対する最終検査 (禁止項目の混入チェック)
FORBIDDEN_PATTERNS = [
    (r"\d+(\.\d+)?\s*倍", "オッズ生値 (倍)"),
    (r"オッズ", "オッズへの言及"),
    (r"補正タイム", "TARGET 補正タイム"),
    (r"\bZI\b", "TARGET ZI 指数"),
    (r"PCI", "PCI 指数"),
    # ウッドは直前がカタカナなら馬名の一部 (ハリウッドメモリー/ナリノシャーウッド等) とみなし除外
    (r"(坂路|(?<![ァ-ヶー])ウッド|CW|美浦W|栗東坂路)", "調教タイム/調教施設"),
    (r"調教タイム", "調教タイム"),
    (r"馬体重", "ライブ馬体重"),
    (r"上が?り3F", "上り3Fタイム"),
    (r"(xROI|EV=|EV ?[:：])", "EV/xROI"),
    (r"JV-?Link", "JV-Link 由来データ"),
]
# 免責文そのものが「調教タイム/オッズは載せていない」と書くため、検査前に除去する
# 定型文 (これ以外の文脈で語が出たら NG 扱い)。
AUDIT_EXEMPT = [
    "本記事は JRA-VAN の投稿ガイドラインに従い、調教タイム・オッズ・払戻金・外部指数等の\n"
    "  データそのものは掲載していません。",
]


def scrub_note(day: dict) -> dict:
    """公開ペイロードに記事固有の追加スクラブを掛ける (破壊的)。"""
    build_site.scrub_public(day)  # 冪等。site/data 側の除外リストをそのまま再適用
    for r in day.get("races", []):
        cw = r.get("cowork") or {}
        # T-10 定型文にオッズ生値 (「単勝 5.9倍」「[T-10オッズ ...]」) が入るので丸ごと落とす
        cw.pop("race_reason", None)
        cw.pop("race_nature", None)
        for b in cw.get("bets") or []:
            b.pop("reason", None)
        for a in cw.get("advisor") or []:
            if BLOCKED_WHY.search(a.get("comment") or ""):
                a["comment"] = ""
        for h in r.get("horses", []) or []:
            why = []
            for w in h.get("why") or []:
                if BLOCKED_WHY.search(f"{w.get('feat')} {w.get('label')}"):
                    continue
                why.append({"label": w.get("label"), "contrib": w.get("contrib")})
            h["why"] = why
    return day


def audit_markdown(text: str) -> list[str]:
    """生成 md に禁止項目が混入していないか機械検査。所見のリストを返す。"""
    for ex in AUDIT_EXEMPT:
        text = text.replace(ex, "")
    found = []
    for pat, name in FORBIDDEN_PATTERNS:
        for m in re.finditer(pat, text):
            s = max(0, m.start() - 40)
            ctx = text[s:m.end() + 40].replace("\n", " ")
            found.append(f"{name}: ...{ctx}...")
    return found


# ------------------------------------------------------------------ note 用の表変換
# note の本文エディタは Markdown 表を解釈しない (パイプ行がそのまま出る)。
# note.com/moctty_note/n/ndc285995bd76 の 4 手法 (TeX/Gist/Excel貼付/画像) のうち、
# 生成スクリプトから自動で出せて note 内完結・後から編集可・本文検索可なのは
# TeX(KaTeX) だけなので、完成 md のパイプ表を \begin{array} に後段変換する。
# (Gist は表の数だけ外部 gist が要る / Excel・画像は自動生成できず検索も効かない)
_TEX_ESC = {"\\": "", "{": r"\{", "}": r"\}", "$": r"\$", "&": r"\&",
            "#": r"\#", "_": r"\_", "%": r"\%", "^": r"\^{}", "~": r"\~{}"}


def _tex_cell(s: str) -> str:
    """1 セルを KaTeX の text ノードにする。**強調** は \\textbf に写す。"""
    s = s.strip()
    if not s:
        return r"\text{}"
    esc = "".join(_TEX_ESC.get(ch, ch) for ch in s)
    out = []
    for i, part in enumerate(re.split(r"\*\*(.+?)\*\*", esc)):
        if not part:
            continue
        out.append(rf"\textbf{{{part}}}" if i % 2 else rf"\text{{{part}}}")
    return "".join(out) or r"\text{}"


def _split_row(line: str) -> list[str]:
    return [c for c in line.strip().strip("|").split("|")]


def _is_sep(line: str) -> bool:
    cells = _split_row(line)
    return bool(cells) and all(re.fullmatch(r"\s*:?-{2,}:?\s*", c) for c in cells)


def _iter_tables(lines: list[str]):
    """md の行列から (開始index, 終了index, ヘッダ, 区切り行, 本文行) を返す。"""
    i = 0
    while i < len(lines):
        if (lines[i].lstrip().startswith("|") and i + 1 < len(lines)
                and _is_sep(lines[i + 1])):
            head = _split_row(lines[i])
            sep = _split_row(lines[i + 1])
            body, j = [], i + 2
            while j < len(lines) and lines[j].lstrip().startswith("|"):
                body.append(_split_row(lines[j]))
                j += 1
            yield i, j, head, sep, body
            i = j
        else:
            i += 1


# 箇条書き化したとき、行頭の見出しにまとめる列 (先頭から連続する分だけ)
_LEAD_COLS = {"印", "馬番", "馬名", "AI 順位", "R", "レース", "券種", "買い目", "項目"}


def md_tables_to_text(text: str) -> str:
    """パイプ表を箇条書きテキストに落とす (note に貼ると表は必ず潰されるため)。

    2 列表は「- **見出し** 値」、3 列以上は 1 行目のセルを太字の見出しにして
    残りを「列名 値」の羅列にする。情報は 1 つも落とさない。
    """
    lines = text.split("\n")
    repl = {}
    for s, e, head, _sep, body in _iter_tables(lines):
        # 行の「見出し」にする先頭列群 (馬を特定する列・券種と買い目 など)
        n_lead = 0
        for h in head:
            if n_lead >= 3 or h.strip() not in _LEAD_COLS:
                break
            n_lead += 1
        n_lead = max(1, n_lead)
        out = []
        for b in body:
            b = (b + [""] * len(head))[:len(head)]
            if len(head) == 2:
                out.append(f"- {b[0]}　{b[1]}" if b[0].strip().startswith("**")
                           else f"- **{b[0].strip()}**　{b[1].strip()}")
                continue
            lead = " ".join(c.strip().strip("*") for c in b[:n_lead] if c.strip())
            rest = [f"{h.strip()} {c.strip()}"
                    for h, c in zip(head[n_lead:], b[n_lead:])
                    if c.strip() and c.strip() != "—"]
            out.append(f"- **{lead or '—'}**　" + "・".join(rest))
        repl[s] = (e, out)
    if not repl:
        return text
    res, i = [], 0
    while i < len(lines):
        if i in repl:
            e, out = repl[i]
            res += out
            i = e
            continue
        res.append(lines[i])
        i += 1
    return "\n".join(res)


def md_tables_to_katex(text: str) -> str:
    """md 中のパイプ表を KaTeX の array ブロックに置換する (表以外は無改変)。"""
    lines = text.split("\n")
    out, i = [], 0
    while i < len(lines):
        if (lines[i].lstrip().startswith("|") and i + 1 < len(lines)
                and _is_sep(lines[i + 1])):
            head = _split_row(lines[i])
            aligns = []
            for c in _split_row(lines[i + 1]):
                c = c.strip()
                aligns.append("c" if c.startswith(":") and c.endswith(":")
                              else "r" if c.endswith(":") else "l")
            body, j = [], i + 2
            while j < len(lines) and lines[j].lstrip().startswith("|"):
                body.append(_split_row(lines[j]))
                j += 1
            # note の数式は「1 段落 = 1 ブロック」。複数行で貼ると $$ や \begin{array} が
            # 別々の段落に割れて、ただの文字列として出る (2026-08-08 プレビューで実証)。
            # そのため改行を一切入れず 1 行に畳む。行区切りは \\ ではなく \cr
            # (note.com/masato_powerup/n/n1e28ba44cbab)。罫線は外枠込みの全グリッド、
            # 行間は \arraystretch。編集画面では描画されず、プレビュー/公開後に出る。
            spec = "|" + "|".join(aligns[:len(head)]) + "|"
            rows = [" & ".join(_tex_cell(c) for c in head) + r" \cr \hline"]
            for b in body:
                b = (b + [""] * len(head))[:len(head)]
                rows.append(" & ".join(_tex_cell(c) for c in b) + r" \cr \hline")
            # note の記法は 2 種類。ディスプレイ数式 ($$ 単独行 → 中身 → $$) は
            # エディタで $$ を打って Enter した時だけブロック化し、同じ文字列を
            # 貼り付けてもただの文字列になる (2026-08-08 プレビューで実証)。
            # 貼り付けで効くのはインライン数式 = $${ 中身 }$$ (波カッコ必須)。
            out.append("$${" + r"\newcommand{\arraystretch}{1.5}"
                       + rf"\begin{{array}}{{{spec}}}\hline "
                       + " ".join(rows) + r"\end{array}" + "}$$")
            i = j
            continue
        out.append(lines[i])
        i += 1
    return "\n".join(out)


# ------------------------------------------------------------------ note 貼付用 HTML
# .md をエディタから素のテキストとしてコピペしても、note は表にできない
# (数式ブロックは $$ を打って Enter した時だけ生成され、貼り付けでは作られない)。
# 一方 note は「表計算ソフトからのコピペ」を表として受け取れる = クリップボードの
# text/html を解釈している。よってブラウザで開いて丸ごとコピーできる HTML を出す。
# これなら見出し・太字・表・リストが 1 回の貼り付けでそのまま入る。
HTML_HEAD = """<meta charset="utf-8">
<title>{title}</title>
<style>
 body{{font-family:"Hiragino Sans","Yu Gothic",sans-serif;line-height:1.9;
      max-width:760px;margin:2rem auto;padding:0 1rem;
      color:#111;background:#fff}}   /* ブラウザのダークモードでも白地固定 */
 table{{border-collapse:collapse;margin:1rem 0;width:100%}}
 th,td{{border:1px solid #999;padding:6px 10px;font-size:14px;white-space:nowrap}}
 th{{background:#f2f2f2}}
 blockquote{{border-left:4px solid #ccc;margin:1rem 0;padding:.5rem 1rem;background:#f7f9fb}}
 h1{{font-size:1.6rem}} h2{{font-size:1.3rem;margin-top:2rem}} h3{{font-size:1.1rem}}
 hr{{border:0;border-top:1px solid #ddd;margin:2rem 0}}
</style>
"""


def md_tables_to_xlsx(md: str, path) -> int:
    """記事内の全パイプ表を xlsx に書き出す。**同じ列構成の表は 1 枚に集約する**。

    note で表として通る唯一の経路が「デスクトップ版 Excel からのコピペ」
    (貼ると画像の表になる。Excel Online / Google スプレッドシートは不可)。
    貼り付け回数 = シート数なので、全レースで列が共通する印テーブル・買い目は
    先頭に R 列を足して 1 枚にまとめ、5 回前後で記事が埋まるようにする。
    シートを開いて Ctrl+A → Ctrl+C → note に貼る、をシートの数だけ繰り返す。
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="F2F2F2")

    lines = md.split("\n")
    heads, cur = {}, "表"
    for n, ln in enumerate(lines):
        if ln.startswith("#"):
            cur = ln.lstrip("#").strip()
        heads[n] = cur

    # 同じ列構成の表をまとめる。R 見出しの下にある表には R 列を足す。
    groups: dict[tuple, dict] = {}
    order: list[tuple] = []
    for s, _e, head, _sep, body in _iter_tables(lines):
        key = tuple(h.strip() for h in head)
        ctx = heads.get(s, "表")
        rno = re.match(r"\s*(\d+)R", ctx)
        if key not in groups:
            groups[key] = {"head": list(key), "rows": [], "ctx": ctx, "n": 0}
            order.append(key)
        g = groups[key]
        g["n"] += 1
        for b in body:
            b = [c.strip().strip("*") for c in (b + [""] * len(key))[:len(key)]]
            g["rows"].append(([f"{rno.group(1)}R"] if rno else [""]) + b)

    wb = Workbook()
    wb.remove(wb.active)
    used = set()
    for key in order:
        g = groups[key]
        multi = g["n"] > 1          # 複数レースをまとめた表だけ R 列を出す
        head = (["R"] if multi else []) + g["head"]
        rows = [(r if multi else r[1:]) for r in g["rows"]]
        base = "印一覧（全R）" if multi and "馬名" in key else \
               "買い目一覧（全R）" if multi else \
               re.sub(r"[\[\]:*?/\\]", "", g["ctx"])[:26] or "表"
        name, k = base, 1
        while name in used:
            k += 1
            name = f"{base[:26]}_{k}"
        used.add(name)
        ws = wb.create_sheet(name)
        ws.freeze_panes = "A2"
        for j, c in enumerate(head, 1):
            cell = ws.cell(1, j, c)
            cell.font = Font(bold=True)
            cell.fill = head_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        for i, b in enumerate(rows, 2):
            for j, c in enumerate(b, 1):
                cell = ws.cell(i, j, c)
                cell.border = border
        for j in range(1, len(head) + 1):
            w = max([len(str(head[j - 1]))]
                    + [len(str(b[j - 1])) for b in rows if j - 1 < len(b)])
            ws.column_dimensions[ws.cell(1, j).column_letter].width = min(30, w * 1.6 + 2)
    wb.save(path)
    return len(wb.sheetnames)


# ------------------------------------------------------------------ 表 → 画像
# note が構造を保ったまま受け取れるのは画像だけ (HTML の <table> は貼付時に
# 1 段落へ潰される / 数式ブロックは打鍵でしか作れない)。デスクトップ Excel からの
# コピペが表になるのも、Excel がクリップボードに画像を載せているから。
# ならば最初から表を画像にして HTML に埋めれば、ページ全体を 1 回コピーするだけで
# 本文 + 表が同時に運べる。
_FONT_R = "C:/Windows/Fonts/YuGothM.ttc"
_FONT_B = "C:/Windows/Fonts/YuGothB.ttc"


def table_png(head: list[str], body: list[list[str]], scale: int = 2) -> bytes:
    """表を PNG にする。scale は解像度倍率 (2 = Retina 相当)。"""
    from PIL import Image, ImageDraw, ImageFont

    fs = 15 * scale
    pad = 10 * scale
    fr = ImageFont.truetype(_FONT_R, fs)
    fb = ImageFont.truetype(_FONT_B, fs)
    rows = [[c.strip().strip("*") for c in r] for r in body]
    head = [c.strip().strip("*") for c in head]
    ncol = len(head)
    rows = [(r + [""] * ncol)[:ncol] for r in rows]

    probe = ImageDraw.Draw(Image.new("RGB", (1, 1)))

    def w(t, f):
        return probe.textlength(t, font=f)

    widths = [int(max([w(head[j], fb)] + [w(r[j], fr) for r in rows]) + pad * 2)
              for j in range(ncol)]
    rh = fs + pad * 2
    W, H = sum(widths) + 1, rh * (len(rows) + 1) + 1

    img = Image.new("RGB", (W, H), "white")
    d = ImageDraw.Draw(img)
    d.rectangle([0, 0, W - 1, rh], fill="#F2F2F2")
    y = 0
    for i, r in enumerate([head] + rows):
        x = 0
        f = fb if i == 0 else fr
        for j, c in enumerate(r):
            d.text((x + pad, y + pad), c, font=f, fill="#111")
            x += widths[j]
        y += rh
    for j in range(ncol + 1):                       # 縦罫線
        x = sum(widths[:j])
        d.line([(x, 0), (x, H - 1)], fill="#999", width=scale)
    for i in range(len(rows) + 2):                  # 横罫線
        d.line([(0, i * rh), (W - 1, i * rh)], fill="#999", width=scale)

    buf = io.BytesIO()
    img.resize((W // scale, H // scale), Image.LANCZOS).save(buf, "PNG")
    return buf.getvalue()


def md_tables_to_img_html(md: str, title: str) -> str:
    """パイプ表を PNG (data URI) に差し替えた HTML。ページ丸ごと 1 回コピー用。"""
    lines = md.split("\n")
    repl = {}
    for s, e, head, _sep, body in _iter_tables(lines):
        png = table_png(head, body)
        b64 = base64.b64encode(png).decode()
        repl[s] = (e, f'<p><img src="data:image/png;base64,{b64}"></p>')
    res, i = [], 0
    while i < len(lines):
        if i in repl:
            e, tag = repl[i]
            res.append(tag)
            i = e
            continue
        res.append(lines[i])
        i += 1
    return md_to_html("\n".join(res), title)


def md_to_html(md: str, title: str) -> str:
    """記事 md を、ブラウザで開いて丸ごとコピーするための HTML にする。"""
    import markdown2
    body = markdown2.markdown(md, extras=["tables", "cuddled-lists"])
    return HTML_HEAD.format(title=title) + body


# ------------------------------------------------------------------ 実績集計
def wilson(k: int, n: int, z: float = 1.96) -> tuple[float, float]:
    if not n:
        return (0.0, 0.0)
    p = k / n
    d = 1 + z * z / n
    c = (p + z * z / (2 * n)) / d
    h = z * math.sqrt(p * (1 - p) / n + z * z / (4 * n * n)) / d
    return (round(100 * (c - h), 1), round(100 * (c + h), 1))


def mark_stats() -> dict:
    """site/data の全日次ペイロードから、実戦での印の成績を Wilson CI 付きで集計。"""
    files = sorted(f for f in glob.glob(str(SITE_DATA / "*.json"))
                   if re.fullmatch(r"\d{8}\.json", os.path.basename(f)))
    st = collections.defaultdict(lambda: [0, 0, 0])
    n_race = 0
    top1_in5 = 0
    dates = []
    for f in files:
        day = json.load(open(f, encoding="utf-8"))
        dates.append(day["date"])
        for r in day.get("races", []):
            res = r.get("result") or {}
            if not res.get("top3"):
                continue
            n_race += 1
            top3 = set(res["top3"])
            first = res["top3"][0]
            marked = set()
            for h in r["horses"]:
                m = h.get("mark") or ""
                if not m:
                    continue
                marked.add(h["umaban"])
                s = st[m]
                s[0] += 1
                s[1] += int(h["umaban"] == first)
                s[2] += int(h["umaban"] in top3)
            if first in marked:
                top1_in5 += 1
    return {"marks": dict(st), "n_race": n_race, "top1_in5": top1_in5,
            "date_from": min(dates), "date_to": max(dates), "n_days": len(dates)}


def day_settlement(date: str) -> dict:
    """cowork_results.json から当日の自分の買い目の決済結果を取り出す。"""
    p = ROOT / "data" / "cowork_results.json"
    d = json.load(open(p, encoding="utf-8"))
    bets = [b for b in d.get("bets", []) if b.get("date") == date]
    races = {(r["場所"], int(r["R"])): r for r in d.get("races", [])
             if r.get("date") == date}
    by_type = collections.defaultdict(lambda: {"n": 0, "bet": 0, "ret": 0, "hit": 0})
    for b in bets:
        t = by_type[b["馬券種"]]
        t["n"] += 1
        t["bet"] += b["購入額"]
        t["ret"] += b["払戻"]
        t["hit"] += int(b["払戻"] > 0)
    tot_bet = sum(b["購入額"] for b in bets)
    tot_ret = sum(b["払戻"] for b in bets)
    all_dates = sorted({r["date"] for r in d.get("races", []) if r.get("date")})
    d0, d1 = all_dates[0], all_dates[-1]
    months = max(1, round((int(d1[:6]) - int(d0[:6])) % 100 + 1
                          if d0[:4] == d1[:4] else 1))
    months = (int(d1[:4]) - int(d0[:4])) * 12 + int(d1[4:6]) - int(d0[4:6])
    return {"bets": bets, "races": races, "by_type": dict(by_type),
            "bet": tot_bet, "ret": tot_ret,
            "roi": round(100 * tot_ret / tot_bet, 1) if tot_bet else 0.0,
            "total": d.get("total", {}), "type_agg": d.get("by_type", {}),
            "range": (f"{d0[:4]}-{d0[4:6]}-{d0[6:]}", f"{d1[:4]}-{d1[4:6]}-{d1[6:]}"),
            "months": max(1, months)}


# ------------------------------------------------------------------ 無料パート
FREE_INTRO = """> この記事の**冒頭（無料部分）に、これまでの成績を勝ち負け両方そのまま載せています**。
> 買う前に必ず読んでください。読んだうえで「これは自分向けじゃない」と思ったら、買わないのが正解です。

## この記事は何か

JRA 中央競馬の全レースを、自作の機械学習モデル（LightGBM の LambdaRank + Plackett-Luce
確率モデル）で処理して、**◎〇▲△△の印と、その馬が勝つ確率・3着以内に入る確率**を出しています。
競馬新聞の予想家の記事ではなく、**モデルの出力をそのまま人間の言葉に翻訳したもの**です。

書いているのは個人です。プロの予想会社でも、投資助言業者でもありません。
"""


def free_part(ms: dict, res: dict, date_disp: str) -> str:
    tot = res["total"]
    ta = res["type_agg"]
    d0, d1 = res["range"]
    L = [FREE_INTRO, "## 先に、負けている数字から出します\n"]
    L.append(f"{d0} 〜 {d1} の実運用（対象 {tot['races']:,} レース、"
             f"実際に賭けたのは {tot['bet_count']:,} 点）の通算成績です。"
             "見送ったレースも分母に入れて、外れも全部そのまま数えています。\n")
    L.append("| 項目 | 実績 |")
    L.append("|---|---|")
    L.append(f"| 総投資 | {tot['bet']:,} 円 |")
    L.append(f"| 総払戻 | {tot['ret']:,} 円 |")
    L.append(f"| **収支** | **{tot['pnl']:+,} 円** |")
    L.append(f"| **回収率（ROI）** | **{tot['roi']}%** |")
    L.append(f"| 的中率（点数ベース） | {tot['hit_rate']}%（{tot['hit_count']:,} / {tot['bet_count']:,} 点） |")
    L.append("")
    L.append(f"**マイナス {abs(tot['pnl'])/10000:.0f} 万円です。** "
             f"約 {res['months']} か月で、投資額の {100 - tot['roi']:.0f}% が消えています。\n")

    L.append("### 券種別（95% 信頼区間つき）\n")
    L.append("| 券種 | 賭けた点数 | 的中率 [95%CI] | 回収率 [95%CI] | 判定 |")
    L.append("|---|---|---|---|---|")
    verdict_ja = {"below_takeout": "**控除率を明確に下回る**", "inconclusive": "判断不能（誤差が大きい）",
                  "above_takeout": "控除率を上回る"}
    for name in ("単勝", "複勝", "ワイド", "馬連", "馬単"):
        t = ta.get(name)
        if not t:
            continue
        hc, rc = t["hit_ci95"], t["roi_ci95"]
        L.append(f"| {name} | {t['races']} | {t['hit_rate']}% [{hc[0]}–{hc[1]}] | "
                 f"{t['roi']}% [{rc[0]}–{rc[1]}] | {verdict_ja.get(t['roi_verdict'], t['roi_verdict'])} |")
    L.append("")
    L.append("信頼区間を見てほしいのですが、**「利益が出ている」と言える券種はひとつもありません。**\n"
             "単勝の回収率 89.5% は上限が 126% まで伸びますが、これは「たまたま良かった可能性も、"
             "もっと悪い可能性も否定できない」という意味であって、勝っている証拠ではありません。\n"
             "馬単だけは信頼区間の上限が控除率にすら届かず、**明確に負ける買い方**と判定されたので、"
             "現在は生成ロジックから外してあります。\n")

    L.append("### 印そのものの精度\n")
    m = ms["marks"]
    L.append(f"同じ期間（{ms['n_days']} 開催日 / {ms['n_race']:,} レース）で、"
             "印がどれだけ当たっているか。こちらは馬券の買い方と関係なく、モデルの素の精度です。\n")
    L.append("| 印 | 対象 | 1着率 [95%CI] | 複勝圏（3着以内）率 [95%CI] |")
    L.append("|---|---|---|---|")
    for k in ("◎", "〇", "▲", "△"):
        if k not in m:
            continue
        n, w, t3 = m[k]
        wc, tc = wilson(w, n), wilson(t3, n)
        L.append(f"| {k} | {n:,} 頭 | {100*w/n:.1f}% [{wc[0]}–{wc[1]}] | "
                 f"{100*t3/n:.1f}% [{tc[0]}–{tc[1]}] |")
    n5 = ms["n_race"]
    c5 = wilson(ms["top1_in5"], n5)
    L.append(f"| （印 5 頭のどれかが 1 着） | {n5:,} R | "
             f"{100*ms['top1_in5']/n5:.1f}% [{c5[0]}–{c5[1]}] | — |")
    L.append("")
    L.append("**ここも正直に書きます。** 過去データでの検証（2024–2025 年、6,878 レース）では "
             "◎の複勝圏率は 61.7% でしたが、**実運用ではまだ 52% 前後までしか出ていません。**\n"
             "約 10 ポイントの差があります。標本が 1,073 レースなので偶然のブレの範囲にぎりぎり入りますが、"
             "本番環境と検証環境の差（特徴量の欠損の出方など）が原因の可能性も潰しきれていないので、"
             "現在も監視中です。**この差を隠したまま売るのは違うと思うので書いておきます。**\n")

    L.append("### 儲かる商品ではありません\n")
    L.append("JRA の控除率は券種によって 20〜30% です。つまり、参加者全体の回収率は"
             "構造的に 70〜80% に固定されていて、**予想の精度をいくら上げても、"
             "その壁の内側でしか動けません。**\n"
             "私はこれまでに、予測モデルの改良・特徴量の追加・期待値による買い目の絞り込みなど、"
             "20 以上の方向で改善を試して、**どれも控除率の壁を超えられませんでした**"
             "（超えたように見えたものは、後で検証すると全部たまたまでした）。\n")
    L.append("だからこの記事が提供できるのは、**「勝てる買い目」ではなく、"
             "「同じ負けるなら、どこがいちばんマシか」の材料**です。\n")
    L.append("- **向いている人**：自分で買い目を組む人。AI が何をどう評価したか（確率・レベル・"
             "根拠）を見て、自分の見立てと突き合わせたい人。数字で判断したい人。\n"
             "- **向いていない人**：勝ちたい人。回収率 100% 超えを期待する人。"
             "「この馬を買え」と言い切ってほしい人。\n")
    L.append("後者の方は買わないでください。期待に応えられません。\n")
    return "\n".join(L)


HOWTO = """## 記事の読み方（各レース共通）

| 項目 | 意味 |
|---|---|
| 印 ◎〇▲△△ | モデルの評価順位そのもの。◎=1 位、〇=2 位、▲=3 位、△=4・5 位。人気とは無関係に付きます |
| 勝つ確率 / 3着内確率 | Plackett-Luce モデルで出した確率を、過去データで較正（キャリブレーション）した値。レース内の全馬の「勝つ確率」を足すと 100% になります |
| 馬レベル | その馬の近走成績（着順・人気などの公表情報のみ）を 0〜100 に正規化した独自指標。同じクラス内での比較用 |
| メンバーレベル | 出走馬の上位 3 頭の馬レベル平均が、**同じクラスの過去のレースと比べて**どの位置か。G3 でも「低調」と出ることがありますが、それは「G3 の平均と比べて」という意味です |
| 妙味 | 人気（＝市場の評価）に対して、モデルの確率が高いか低いかの 5 段階。S ほど「人気の割にモデルは買っている」、「罠」は逆に「人気ほどには評価していない」 |
| AI vs 市場 | 同じことを 3 段階で表したもの |
| 混戦度 | レース全体の確率の散らばり。堅いほど上位数頭に確率が集中しています |

**確率は「当たる保証」ではありません。** たとえば 3着内確率 60% の◎は、10 回走れば 4 回は
馬券圏外に飛ぶ、という意味です。上の無料パートに書いたとおり、実運用での◎の 3 着内率は
まだ 52% 前後です。

なお、**買い目に 1 点ごとの金額は載せていません。** 資金量もリスク許容度も人によって
違うのに、「1 点 3,000 円」と書いた瞬間それが正解のように見えてしまうからです。
1 日の合計収支だけは、成績開示のため記事の最後に出しています。
"""


# ------------------------------------------------------------------ レース記述
UMAMI_JA = {"S": "S（かなり割安）", "A": "A（割安）", "B": "B（やや割安）",
            "C": "C（ほぼ適正）", "罠": "罠（割高・危険）"}
LEVEL_JA = {"S": "S", "A": "A", "B": "B", "C": "C", "D": "D"}
VS_JA = {"under": "市場より高評価", "over": "市場より低評価",
         "fair": "市場と同評価", "unknown": "—"}


def _pct_word(p):
    if p is None:
        return "不明"
    if p >= 0.8:
        return "かなり割れている（荒れ寄り）"
    if p >= 0.6:
        return "やや割れている"
    if p >= 0.4:
        return "標準的"
    if p >= 0.2:
        return "やや堅い"
    return "かなり堅い"


def _fmt_pct(x):
    return "—" if x is None else f"{100*x:.1f}%"


def race_heading(r: dict) -> str:
    name = r.get("race_name") or ""
    klass = (r.get("klass") or "").strip()
    tail = f"{name}（{klass}）" if name and klass else (name or klass or "")
    return f"{r['rno']}R {tail} {r['course']} / {r['field_size']}頭"


def marks_table(r: dict) -> str:
    rows = ["| 印 | 馬番 | 馬名 | 騎手 | 人気 | 勝つ確率 | 3着内確率 | 馬レベル | 妙味 | AI vs 市場 |",
            "|---|---|---|---|---|---|---|---|---|---|"]
    marked = [h for h in r["horses"] if h.get("mark")]
    marked.sort(key=lambda h: h.get("ai_rank") or 99)
    for h in marked:
        lv = h.get("level") or {}
        lvs = f"{lv.get('tier','—')}（{lv.get('score','—')}）" if lv else "—"
        umo = h.get("umami") or {}
        ums = UMAMI_JA.get(umo.get("grade"), "—")
        if umo.get("side") and umo.get("grade") not in (None, "罠"):
            ums += f"／{umo['side']}"
        rows.append(
            f"| {h['mark']} | {h['umaban']} | {h['name']} | {h.get('jockey') or '—'} | "
            f"{h.get('ninki') or '—'} | {_fmt_pct(h.get('p_win'))} | {_fmt_pct(h.get('p_sho'))} | "
            f"{lvs} | {ums} | {VS_JA.get(h.get('vs_market'), '—')} |")
    return "\n".join(rows)


def why_line(h: dict) -> str:
    ws = [w for w in (h.get("why") or []) if w.get("label")]
    if not ws:
        return ""
    up = [w["label"] for w in ws if (w.get("contrib") or 0) > 0][:3]
    dn = [w["label"] for w in ws if (w.get("contrib") or 0) < 0][:2]
    parts = []
    if up:
        parts.append("評価を押し上げた要素は " + "・".join(up))
    if dn:
        parts.append("引き下げた要素は " + "・".join(dn))
    return "、".join(parts) + "。" if parts else ""


def race_commentary(r: dict) -> str:
    """レースの見立てを自然文で組み立てる（数値はモデル出力と公知事実のみ）。"""
    j = r.get("judgment") or {}
    conf = r.get("confidence") or {}
    ml = r.get("member_level") or {}
    cp = r.get("class_prior") or {}
    hs = sorted([h for h in r["horses"] if h.get("mark")],
                key=lambda h: h.get("ai_rank") or 99)
    if not hs:
        return "印が付いていません。"
    hon = hs[0]
    S = []

    # 1. レースの骨格
    s = f"混戦度は{_pct_word(j.get('chaos_pct'))}。"
    dom = conf.get("top1_dominance")
    conc = conf.get("top2_concentration")
    if dom is not None:
        if dom >= 0.10:
            s += (f"◎は〇に対して勝つ確率で {100*dom:.1f} ポイント上回っていて、"
                  "上位は割れていない。")
            if conc:
                s += f"◎〇の 2 頭で勝つ確率の {100*conc:.0f}% を占める。"
        elif dom < 0.03:
            s += "◎と〇の力差がほとんどなく、頭を固定しにくい形。"
        else:
            s += (f"◎が勝つ確率は {_fmt_pct(hon.get('p_win'))}。"
                  "〇との差は小さく、頭は絞りきれない。")
    if ml.get("tier"):
        s += f"メンバーレベルは{ml['tier']}（{ml['label']}）"
        if ml.get("pct") is not None:
            if ml["pct"] >= 50:
                s += f"。同じクラスの過去のレースと比べて上位 {100 - ml['pct']}% の水準。"
            else:
                s += f"。同じクラスの過去のレースと比べると下位 {ml['pct']}% の水準。"
        else:
            s += "。"
    elif ml:
        s += "メンバーレベルは判定不能（出走馬の近走データが足りない）。"
    S.append(s)

    # 2. ◎について
    lv = hon.get("level") or {}
    s = f"◎は{hon['umaban']}番{hon['name']}"
    if hon.get("ninki"):
        s += f"（{hon['ninki']}番人気）"
    s += f"。3 着以内に入る確率は {_fmt_pct(hon.get('p_sho'))}"
    if lv.get("tier"):
        s += f"、近走成績から見た馬レベルは {lv['tier']}（100 点満点で {lv['score']}）"
    s += "。"
    if hon.get("style"):
        s += f"脚質は{hon['style']}。"
    w = why_line(hon)
    if w:
        s += w
    S.append(s)

    # 3. 相手構成
    if len(hs) >= 3:
        oth = hs[1:3]
        s = "相手は" + "、".join(
            f"{h['mark']}{h['umaban']}番{h['name']}（3 着内 {_fmt_pct(h.get('p_sho'))}）"
            for h in oth) + "。"
        if cp.get("hon_top3_pct"):
            s += (f"このクラスの過去実績では、◎の 3 着内率が {cp['hon_top3_pct']}%、"
                  f"〇が {cp['tai_top3_pct']}%、▲が {cp['san_top3_pct']}%"
                  f"（{cp['n_samples']:,} レース集計）。")
        S.append(s)

    # 4. 妙味 / 注意点
    vh = [v for v in (j.get("value_horses") or [])][:3]
    if vh:
        marked_u = {h["umaban"] for h in hs}
        parts = []
        for v in vh:
            t = f"{v['umaban']}番{v['horse_name']}"
            if v["umaban"] not in marked_u:  # 印なしの馬は確率を添える
                t += f"（印なし・勝つ確率 {_fmt_pct(v.get('p_win'))}）"
            parts.append(t)
        S.append("モデルが「人気の割に確率が高い」と見ているのは " + "、".join(parts) + "。")
    warns = []
    if (j.get("chaos_pct") or 0) >= 0.8:
        warns.append("混戦度が高く、上位 5 頭で決まらない可能性が相応にある")
    if hon.get("vs_market") == "over":
        warns.append("◎はモデルより市場のほうを高く買っていて、人気先行の可能性がある")
    if (lv.get("score") or 100) < 35:
        warns.append("◎の近走実績が薄く、レベル評価は低い")
    trap = [h for h in hs if (h.get("umami") or {}).get("grade") == "罠"]
    if trap:
        warns.append("印のうち " + "・".join(f"{h['umaban']}番{h['name']}" for h in trap[:2])
                     + " は人気に対して確率が見合っていない")
    if (r.get("klass") or "") == "新馬":
        warns.append("新馬戦は過去走がないぶん、モデルの判別力が落ちる")
    if warns:
        S.append("注意点：" + "。".join(warns) + "。")
    return "\n".join(S)


def advisor_block(r: dict) -> str:
    cw = r.get("cowork") or {}
    adv = [a for a in (cw.get("advisor") or []) if a.get("comment")]
    if not adv:
        return ""
    tag_ja = {"軸": "軸", "妙味": "妙味", "穴": "穴", "罠": "罠", None: "—"}
    # open 必須: 閉じた <details> の中身はブラウザの全選択コピーに含まれず、
    # note 貼付で短評が丸ごと欠落する (2026-08-15)
    L = ["<details open><summary>AI の各馬短評</summary>", "",
         "※ 短評で取り上げる馬は、印（確率順の上位 5 頭）とは別に選んでいます。"
         "そのため印の付いていない馬が入ることがあります。", ""]
    for a in adv:
        tag = tag_ja.get(a.get("tag"), a.get("tag") or "—")
        L.append(f"- **{a['umaban']}番 {a['horse_name']}**（{tag}）… {a['comment']}")
    L += ["", "</details>"]
    return "\n".join(L)


def load_tact(date: str) -> dict:
    """レース前の買い目を bundle から直接組む (2026-08-08)。

    site/data/{date}.json は scrub_public() 済みで、結果が出るまで cowork/tact の
    bets が落ちている (買い目はサイト非公開 = note 専売の仕様)。記事はレース前に
    売るものなので、公開ペイロードではなく bundle から TACT (topdown エンジン) を
    その場で組み直す。理由文のオッズ除去は build_tact 側で済み、最終 md は
    audit_markdown() が再検査する。
    """
    p = ROOT / "reports" / "cowork_input" / f"{date}_bundle.json"
    if not p.exists():
        return {}
    b = json.load(open(p, encoding="utf-8"))
    races = b.get("races") if isinstance(b, dict) else b
    out = {}
    for race in races or []:
        rid = str(race.get("race_id") or "")
        t = build_site.build_tact(race)
        if rid and t and t.get("bets"):
            out[rid] = {"version": t.get("version", ""), "bets": t["bets"]}
    return out


def bets_block(r: dict, settle: dict) -> str:
    cw = r.get("cowork") or {}
    bets = cw.get("bets") or []
    key = (r["place"], r["rno"])
    row = settle["races"].get(key)
    if not bets:
        # レース前記事: 決済前は cowork/tact とも scrub 済み → bundle 由来の TACT を出す
        tb = TACT_BETS.get(str(r.get("race_id") or ""))
        if tb:
            L = [f"**買い目（BetAI TACT v{tb['version']}）**", "",
                 "| 券種 | 買い目 | 根拠 |", "|---|---|---|"]
            for b in tb["bets"]:
                L.append(f"| {b['type']} | {b['selection']} | {b.get('reason','')} |")
            L += ["", "※ 金額は書きません（資金量もリスク許容度も人によって違うため）。"
                  "点数配分だけ真似して、額はご自身で決めてください。"]
            return "\n".join(L)
        j = r.get("judgment") or {}
        hint = j.get("kenshu_hint") or j.get("headline") or ""
        s = "**この日の私の買い目**：なし（見送り）。"
        s += ("1 日の予算配分と信頼度の枠（勝負／準勝負／消化）で組んだ結果、"
              "このレースには枠が回りませんでした。")
        if hint:
            s += f"モデル側の買い方ヒントは「{hint}」でした（参考値で、拘束ではありません）。"
        return s
    if SHOW_AMOUNTS:
        L = ["**この日の私の買い目**", "", "| 券種 | 買い目 | 金額 |", "|---|---|---|"]
        for b in bets:
            L.append(f"| {b['type']} | {b['selection']} | {int(b.get('amount') or 0):,} 円 |")
    else:
        # 1 点ごとの金額は出さない (資金量もリスク許容度も人によって違うため。
        # docs/x_post_stock.md「なぜ金額を公開しないか」と同じ方針)
        L = ["**この日の私の買い目**", "", "| 券種 | 買い目 |", "|---|---|"]
        for b in bets:
            L.append(f"| {b['type']} | {b['selection']} |")
    if row:
        L.append("")
        n_hit = row["的中点数"]
        if SHOW_RACE_PAYOUT:
            L.append(f"→ 結果：{row['点数']} 点中 {n_hit} 点的中、"
                     f"投資 {row['総投資']:,} 円 / 回収 {row['総払戻']:,} 円"
                     f"（{row['総払戻'] - row['総投資']:+,} 円）")
        else:
            L.append(f"→ 結果：{row['点数']} 点中 {n_hit} 点的中"
                     "（収支は記事末の 1 日合計にまとめています）")
    return "\n".join(L)


def result_block(r: dict) -> str:
    res = r.get("result") or {}
    top3 = res.get("top3") or []
    if not top3:
        return ""
    name = {h["umaban"]: h for h in r["horses"]}
    mk = {h["umaban"]: (h.get("mark") or "") for h in r["horses"]}
    parts = []
    for i, u in enumerate(top3, 1):
        h = name.get(u, {})
        m = mk.get(u) or "無印"
        parts.append(f"{i}着 {u}番 {h.get('name','?')}（{m}）")
    hon = next((h for h in r["horses"] if h.get("mark") == "◎"), None)
    verdict = ""
    if hon:
        verdict = ("◎的中" if hon["umaban"] in set(top3) else "◎は 3 着以内に来ず")
    return f"**結果**：{' / '.join(parts)} → {verdict}"


def race_section(r: dict, settle: dict, show_result: bool) -> str:
    L = [f"### {race_heading(r)}", ""]
    meta = []
    if r.get("start_time"):
        meta.append(f"発走 {r['start_time']}")
    if r.get("baba"):
        meta.append(f"馬場 {r['baba']}")
    if r.get("weather"):
        meta.append(f"天候 {r['weather']}")
    if meta:
        L += [" / ".join(meta), ""]
    L += [marks_table(r), ""]
    L += ["**見立て**", "", race_commentary(r), ""]
    ab = advisor_block(r)
    if ab:
        L += [ab, ""]
    L += [bets_block(r, settle), ""]
    if show_result:
        rb = result_block(r)
        if rb:
            L += [rb, ""]
    L.append("---")
    return "\n".join(L)


# ------------------------------------------------------------------ 重賞
def full_field_table(r: dict) -> str:
    """重賞記事用: 出走全頭のモデル評価 (印 5 頭以外も全部出す)。"""
    rows = ["| AI 順位 | 印 | 馬番 | 馬名 | 騎手 | 人気 | 勝つ確率 | 3着内確率 | 馬レベル | 脚質 |",
            "|---|---|---|---|---|---|---|---|---|---|"]
    for h in sorted(r["horses"], key=lambda x: x.get("ai_rank") or 99):
        lv = h.get("level") or {}
        lvs = f"{lv.get('tier')}（{lv.get('score')}）" if lv.get("tier") else "—"
        rows.append(
            f"| {h.get('ai_rank') or '—'} | {h.get('mark') or ''} | {h['umaban']} | {h['name']} | "
            f"{h.get('jockey') or '—'} | {h.get('ninki') or '—'} | {_fmt_pct(h.get('p_win'))} | "
            f"{_fmt_pct(h.get('p_sho'))} | {lvs} | {h.get('style') or '—'} |")
    return "\n".join(rows)


def grade_section(r: dict, settle: dict, show_result: bool) -> str:
    gs = r.get("grade_scope") or {}
    L = [f"### {race_heading(r)}", ""]
    L += [marks_table(r), ""]
    L += ["**出走全頭のモデル評価**（印の付かなかった馬も含めて全部出します）", "",
          full_field_table(r), ""]
    if gs.get("markdown"):
        # grade_scope は h2 見出しなので、記事内では h4 に下げる
        md = re.sub(r"^## ", "#### ", gs["markdown"], flags=re.M)
        L += ["**このレースの読み**（コース性質・展開・注目馬）", "",
              "※ ここで名前が挙がる馬は、印（確率順の上位 5 頭）とは別枠で、"
              "コース適性や臨戦過程から選んでいます。印と一致しないことがあります。", "",
              md, ""]
    L += ["**モデルの数字での見立て**", "", race_commentary(r), ""]
    ab = advisor_block(r)
    if ab:
        L += [ab, ""]
    L += [bets_block(r, settle), ""]
    if show_result:
        rb = result_block(r)
        if rb:
            L += [rb, ""]
    L.append("---")
    return "\n".join(L)


# ------------------------------------------------------------------ 組み立て
FOOTER = """
## 免責と出典

- この記事は個人が趣味で運用している予測モデルの出力を公開したものです。
  的中や利益を保証するものではありません。馬券の購入は自己判断・自己責任でお願いします。
- 20 歳未満の方は馬券を購入できません。
- 記載の印・確率・馬レベル・妙味グレード・考察は、すべて筆者が自作したモデルによる出力です。
- レース名・出走馬・騎手・着順・人気などの事実情報は、日本中央競馬会（JRA）公表情報および
  株式会社 JRA システムサービスの提供データに基づいて筆者が集計・加工したものです。
- 本記事は JRA-VAN の投稿ガイドラインに従い、調教タイム・オッズ・払戻金・外部指数等の
  データそのものは掲載していません。
"""


# レース単位の回収額を出すか (1点だけ的中した race では配当が逆算できるため、
# 落としたい場合は --no-race-payout。1日/会場単位の合計は常に出す)。
SHOW_RACE_PAYOUT = True
# 1 点ごとの購入金額を出すか (既定は出さない。--show-amounts で出す)
SHOW_AMOUNTS = False
# レース前記事用の TACT 買い目 (race_id -> {version, bets})。main() で load_tact()。
TACT_BETS: dict = {}

COVERAGE_NOTE = ("※ 障害戦など、モデルの学習対象外のレースは掲載していません"
                 "（そのため R 番号が飛ぶことがあります）。")


def paywall(price: str, what: str) -> str:
    return (f"\n---\n\n## ここから有料（{price}）\n\n{what}\n\n"
            "※ note の仕様上、ここから下が購入者のみ閲覧できる範囲です。\n\n---\n")


def weekday_ja(date: str) -> str:
    """20260808 → '土'。土日開催が混ざるので固定文字にしない。"""
    try:
        d = datetime.date(int(date[:4]), int(date[4:6]), int(date[6:8]))
    except ValueError:
        return ""
    return "月火水木金土日"[d.weekday()]


def build_venue(day, place, ms, settle, price, show_result):
    races = [r for r in day["races"] if r["place"] == place]
    races.sort(key=lambda r: r["rno"])
    date_disp = f"{day['date'][:4]}-{day['date'][4:6]}-{day['date'][6:]}"
    wd = weekday_ja(day["date"])
    L = [f"# 【AI 予想】{date_disp}（{wd}）{place} 全{len(races)}R ｜ 印・確率・見立て全公開",
         ""]
    L += [free_part(ms, settle, date_disp)]
    n_grade = sum(1 for r in races if r.get("grade_scope"))
    L += [paywall(price,
                  f"{place}全{len(races)}レース分{'（重賞含む）' if n_grade else ''}"
                  "の印（◎〇▲△△）、"
                  "各馬の勝つ確率と 3 着内確率、馬レベル、妙味グレード、"
                  "AI の各馬短評、レースごとの見立て、そして私自身の買い目"
                  + ("と結果。" if show_result else "（または見送り判断）。"))]
    L += [HOWTO, ""]
    L += [f"## {place} 全{len(races)}R", ""]
    L += [COVERAGE_NOTE, ""]
    L += ["| R | レース | 印◎ | 買い目 |", "|---|---|---|---|"]
    for r in races:
        hon = next((h for h in r["horses"] if h.get("mark") == "◎"), None)
        nb = len((r.get("cowork") or {}).get("bets") or []) or \
            len((TACT_BETS.get(str(r.get("race_id") or "")) or {}).get("bets") or [])
        L.append(f"| {r['rno']} | {(r.get('race_name') or r.get('klass') or '')} {r['course']} | "
                 f"{hon['name'] if hon else '—'} | {f'{nb} 点' if nb else '見送り'} |")
    L += [""]
    for r in races:
        if r.get("grade_scope"):
            L.append(grade_section(r, settle, show_result))
        else:
            L.append(race_section(r, settle, show_result))
    L.append(day_wrap(day, place, settle))
    L.append(FOOTER)
    return "\n".join(L)


def build_grade(day, r, ms, settle, price, show_result):
    date_disp = f"{day['date'][:4]}-{day['date'][4:6]}-{day['date'][6:]}"
    nm = r.get("race_name") or r.get("klass")
    L = [f"# 【AI 予想】{date_disp} {nm}（{r['klass']}）｜ 印・確率・見立て", ""]
    L += [free_part(ms, settle, date_disp)]
    L += [paywall(price, f"{nm} 1 レース分の印（◎〇▲△△）、各馬の勝つ確率と 3 着内確率、"
                         "馬レベル、妙味グレード、コース性質と展開の読み、AI の各馬短評、"
                         "そして私自身の買い目（または見送り判断）と結果。")]
    L += [HOWTO, ""]
    L.append(grade_section(r, settle, show_result))
    L.append(FOOTER)
    return "\n".join(L)


def build_all(day, ms, settle, price, show_result):
    date_disp = f"{day['date'][:4]}-{day['date'][4:6]}-{day['date'][6:]}"
    places = day["places"]
    n = len(day["races"])
    wd = weekday_ja(day["date"])
    n_grade = sum(1 for r in day["races"] if r.get("grade_scope"))
    grade_note = f"（重賞 {n_grade} 鞍を含む）" if n_grade else ""
    L = [f"# 【AI 予想】{date_disp}（{wd}）全{n}R ｜ "
         f"{'・'.join(places)} {len(places)} 場まとめ", ""]
    L += [free_part(ms, settle, date_disp)]
    L += [paywall(price, f"{'・'.join(places)} の全{n}レース分{grade_note}の印、"
                         "各馬の勝つ確率と 3 着内確率、馬レベル、妙味グレード、"
                         "AI の各馬短評、レースごとの見立て、私自身の買い目"
                         + ("と結果。" if show_result else "（または見送り判断）。"))]
    L += [HOWTO, ""]
    L += ["## 目次", ""]
    for p in places:
        rs = [r for r in day["races"] if r["place"] == p]
        L.append(f"- [{p}（{len(rs)}R）](#{PLACE_SLUG.get(p, p)})")
    L += ["", COVERAGE_NOTE, ""]
    for p in places:
        rs = sorted([r for r in day["races"] if r["place"] == p], key=lambda r: r["rno"])
        L += [f"<a id=\"{PLACE_SLUG.get(p, p)}\"></a>", "", f"## {p} 全{len(rs)}R", ""]
        for r in rs:
            if r.get("grade_scope"):
                L.append(grade_section(r, settle, show_result))
            else:
                L.append(race_section(r, settle, show_result))
        L.append(day_wrap(day, p, settle))
    L.append(day_wrap(day, None, settle))
    L.append(FOOTER)
    return "\n".join(L)


def day_wrap(day, place, settle):
    """当日の的中報告（会場単位 / 全体）。自分の買い目の収支のみ。"""
    rows = [v for k, v in settle["races"].items() if place is None or k[0] == place]
    bet = sum(r["総投資"] for r in rows)
    ret = sum(r["総払戻"] for r in rows)
    n = sum(r["点数"] for r in rows)
    hit = sum(r["的中点数"] for r in rows)
    label = place if place else f"{len(day['places'])} 場合計"
    if not bet:
        # 決済行ゼロには 2 通りある: (a) レース前 = 結果未着 (b) 結果は出たが全見送り。
        # レース前記事で「全レース見送り」と出すのは誤り (2026-08-15 ユーザー指摘)
        has_result = any((r.get("result") or {}).get("top3")
                         for r in day["races"] if place is None or r["place"] == place)
        if not has_result:
            return (f"### この日の結果（{label}）\n\n"
                    "レース結果が出次第、ここに追記します。\n")
        return f"### この日の結果（{label}）\n\nこの会場では買い目なし（全レース見送り）。\n"
    marks_hit = 0
    marks_n = 0
    for r in day["races"]:
        if place and r["place"] != place:
            continue
        res = r.get("result") or {}
        hon = next((h for h in r["horses"] if h.get("mark") == "◎"), None)
        if res.get("top3") and hon:
            marks_n += 1
            marks_hit += int(hon["umaban"] in set(res["top3"]))
    L = [f"### この日の結果（{label}）", "",
         f"- 買い目 {n} 点中 {hit} 点的中",
         f"- 投資 {bet:,} 円 / 回収 {ret:,} 円 → **{ret - bet:+,} 円**"
         f"（回収率 {100*ret/bet:.1f}%）"]
    if marks_n:
        L.append(f"- ◎の 3 着以内率：{marks_hit}/{marks_n}（{100*marks_hit/marks_n:.1f}%）")
    L.append("")
    L.append("良い日ではありませんでした。こういう日もそのまま出します。"
             if ret < bet else "この日はプラスで終えられました。ただし短期の結果です。")
    L.append("")
    return "\n".join(L)


# ------------------------------------------------------------------ main
def main():
    global SHOW_RACE_PAYOUT, SHOW_AMOUNTS, TACT_BETS
    date = sys.argv[1] if len(sys.argv) > 1 else "20260802"
    show_result = "--no-result" not in sys.argv
    SHOW_RACE_PAYOUT = "--no-race-payout" not in sys.argv
    SHOW_AMOUNTS = "--show-amounts" in sys.argv
    # レース前記事は公開ペイロードに買い目が無い (scrub_public) → bundle から組む
    if "--no-tact" not in sys.argv:
        TACT_BETS = load_tact(date)
        print(f"TACT 買い目: {len(TACT_BETS)} レース分 "
              f"({sum(len(v['bets']) for v in TACT_BETS.values())} 点)")
    src = SITE_DATA / f"{date}.json"
    if not src.exists():
        sys.exit(f"公開ペイロードが見つかりません: {src}（先に build_site.py を実行）")
    day = json.load(open(src, encoding="utf-8"))
    scrub_note(day)

    ms = mark_stats()
    settle = day_settlement(date)

    out_dir = OUT_ROOT / date
    out_dir.mkdir(parents=True, exist_ok=True)

    # .md の表形式: 既定 KaTeX (--table md で素の Markdown のまま)
    katex = "--table" not in sys.argv or \
        sys.argv[sys.argv.index("--table") + 1:sys.argv.index("--table") + 2] == ["katex"]
    conv = md_tables_to_katex if katex else (lambda t: t)
    print(f"md の表形式: {'KaTeX' if katex else 'Markdown'} / "
          f"note 貼付は .html をブラウザで開いて全選択コピー")

    # 重賞単体パックは廃止 (2026-08-08 ユーザー決定)。商品は会場バラ (¥100) と
    # 全場パック (¥100×会場数、3会場なら¥300) の 2 本立て (2026-08-15 値下げ)。
    # 重賞は所属会場の記事内に grade_section でフル収録されるので、
    # 単体 md は作らない (build_grade は残置・未使用)。
    arts = [(PLACE_SLUG.get(p, p), build_venue(day, p, ms, settle, "¥100", show_result))
            for p in day["places"]]
    arts.append(("all", build_all(day, ms, settle,
                                  f"¥{100 * len(day['places'])}", show_result)))

    written = []
    for slug, raw in arts:
        fp = out_dir / f"{slug}.md"
        fp.write_text(conv(raw), encoding="utf-8")
        written.append(fp)
        # note へはこちらを使う: ブラウザで開いて Ctrl+A → Ctrl+C → note に貼付。
        # 表は note が貼付時に必ず潰す (HTML の <table> も 1 段落に連結される。
        # 2026-08-08 実測) ので、html を作る前に箇条書きテキストへ落としておく。
        title = raw.split("\n")[0].lstrip("# ")
        # ★これが note へ貼るファイル。ブラウザで開いて Ctrl+A → Ctrl+C → 1 回で完成。
        # 表は箇条書きに落としてある (note は貼り付けられた表を必ず潰すため。
        # <table> は 1 段落に連結・KaTeX は打鍵でしかブロック化しない — 2026-08-08 実測)。
        fh = out_dir / f"{slug}.html"
        fh.write_text(md_to_html(md_tables_to_text(raw), title), encoding="utf-8")
        written.append(fh)
        if "--no-img" not in sys.argv:  # 表を PNG 化して埋めた版 (2026-08-09 既定 ON。
            # note 貼付で構造が生きるのは画像のみ。貼付テストで問題なければ ★本命ファイル)
            fi = out_dir / f"{slug}_img.html"
            fi.write_text(md_tables_to_img_html(raw, title), encoding="utf-8")
            written.append(fi)
        if "--xlsx" in sys.argv:     # Excel からシート単位でコピペする版 (既定 OFF)
            fx = out_dir / f"{slug}.xlsx"
            try:
                n_tbl = md_tables_to_xlsx(raw, fx)
            except PermissionError:
                fx = out_dir / f"{slug}_new.xlsx"
                n_tbl = md_tables_to_xlsx(raw, fx)
                print(f"  [warn] {slug}.xlsx は Excel で開かれています → {fx.name}")
            print(f"  {fx.name}: {n_tbl} シート")

    # ---- 禁止項目の機械検査
    lines = [f"JRA-VAN 投稿ガイドライン 混入チェック  ({date})",
             "=" * 60, ""]
    total_hits = 0
    for fp in written:
        text = fp.read_text(encoding="utf-8")
        if fp.suffix == ".html":
            # html は md と同一内容の描画。免責の定型文がタグ挿入で改行位置ごと
            # 変わり AUDIT_EXEMPT に一致しなくなるので、md 側の判定を正とする。
            lines.append(f"[--] {fp.name}  ({len(text):,} 文字) md と同内容のため検査省略")
            continue
        hits = audit_markdown(text)
        total_hits += len(hits)
        lines.append(f"[{'NG' if hits else 'OK'}] {fp.name}  ({len(text):,} 文字)")
        for h in hits[:20]:
            lines.append(f"      - {h}")
    lines += ["", f"検査パターン: {len(FORBIDDEN_PATTERNS)} 種 / 所見 {total_hits} 件", "",
              "検査対象パターン:"]
    lines += [f"  - {n}" for _, n in FORBIDDEN_PATTERNS]
    lines += ["", "自動検査では捕まらない判断事項 (人間の確認が必要):",
              "  1. 妙味グレード(S/A/B/C/罠) と AI vs 市場(under/fair/over) は",
              "     オッズ由来だが 3〜5 段階のラベルで、生値・レンジは出していない。",
              "     公開サイト (build_site.scrub_public) と同じ基準。",
              "  2. 勝つ確率 + 妙味グレードの組み合わせから、単勝オッズのおおまかな",
              "     範囲は推定しうる (ラベル境界の逆算)。site/data と同水準。",
              "  3. レース単位の『回収 N 円』は自分の的中報告だが、1 点のみ的中した",
              "     レースではその組み合わせの配当が逆算できる。落とす場合は",
              f"     --no-race-payout （現在: {'出力あり' if SHOW_RACE_PAYOUT else '出力なし'}）。",
              "  3b. 1 点ごとの購入金額は既定で非表示 (docs/x_post_stock.md の方針に合わせた)。",
              f"      --show-amounts で出せる（現在: {'出力あり' if SHOW_AMOUNTS else '出力なし'}）。",
              "  4. 人気順位は公表事実なので掲載可。ただし事前公開記事では",
              "     発走前オッズに基づく暫定値である点の注記を検討。"]
    rep = "\n".join(lines)
    (out_dir / "_compliance_report.txt").write_text(rep, encoding="utf-8")
    print(rep)


if __name__ == "__main__":
    main()
